# -*- coding: utf-8 -*-
# Version: Python 3.9.5
# Author: TRIX
# Use:大部分不熟悉的内容都能在这里找得到解决方法

# -*-coding:utf8 -*-
#声明python用utf-8编码py文件

#!/usr/bin/python3
#只对 Linux/Unix 用户适用 用来指定本脚本用什么解释器来执行

#版本查看
cmd python --version
#python 版本更新 https://www.cnblogs.com/huny/p/13911721.html

#pip 在终端使用(win+r cmd)
pip --version#查看pip版本
#一串代码更改pip安装默认源 cmd (已设置)
pip config set global.index-url https://pypi.tuna.tsinghua.edu.cn/simple
#清华源镜像安装库
pip install package -i https://pypi.tuna.tsinghua.edu.cn/simple/
pip unistall package
pip list#查看安装的所有模块
#库迁移 cmd
pip freeze > python_module_requirement.txt
#导出python安装的库的名字和版本
pip install -r python_module_requirement.txt #安装所有库
#pip更新
#C:\users\trix\appdata\local\programs\python\python39\python.exe -m pip install --upgrade pip
#pip手动安装
#去https://www.lfd.uci.edu/~gohlke/pythonlibs/#packagename 下载对应的模块 packagename是包名
#模块版本 -cp39-win_amd64.whl

#编辑器默认编码格式设置
#python 3 默认使用utf-8编码
#Pycharm 设置步骤
#进入 file > Settings 在输入框搜索 encoding
#找到 Editor > File encodings 将 IDE Encoding 和 Project Encoding 设置为utf-8

if __name__ == '__main__':#只有本程序运行的时候 if内的语句才会执行 当作为模块导入的时候 不执行if内语句

#打印与格式化输出
print(var)
print(var, end=str)
print(r'str')#不转义字符
import pprint
pprint.pprint(var)#方便阅读的打印模式
pprint.pformat(obj)#返回方便阅读的对象
print('*{index:formatize}'.format(*arg)) #format可以接受不限个参数 位置可以不按顺序 推荐使用format而不是%
"{} {}".format("hello", "world")    # 不设置指定位置 按默认顺序
>>>'hello world'
"{0} {1}".format("hello", "world")  # 设置指定位置
>>>'hello world'
"{1} {0} {1}".format("hello", "world")  # 用索引
>>>'world hello world'
print("网站名 {name}, 地址 {url}".format(name="菜鸟教程", url="www.runoob.com"))#用参数
site={"name": "菜鸟教程", "url": "www.runoob.com"}
print("网站名 {name}, 地址 {url}".format(**site))#用字典
my_list=['菜鸟教程', 'www.runoob.com']
print("网站名 {0[0]}, 地址 {0[1]}".format(my_list))  #用列表  "0" 是必须的
class AssignValue():
    def __init__(self, value):
        self.value=value
my_value=AssignValue(6)
print('value 为: {0.value}'.format(my_value))  #用类 "0" 是可选的

#数字格式化
输入			格式			输出				描述
math.pi	{:.2f}			3.14				#保留小数点后两位
math.pi	{:+.2f}		+3.14			#带符号保留小数点后两位
-1				{:+.2f}		-1.00			#带符号保留小数点后两位
2.71828	{:.0f}			3					#不带小数
5				{:0>2d}		05				#数字补零 (填充左边, 宽度为2)
5				{:>3d}		  5				#空格填充 (填充左边, 宽度为3)
5				{:x<4d}		5xxx				#数字补x (填充右边, 宽度为4)
10			{:x<4d}		10xx				#数字补x (填充右边, 宽度为4)
1000000	{:,}			1,000,000	#以逗号分隔的数字格式
0.25			{:.2%}		25.00%			#百分比格式 保留小数点后两位
1000000	{:.2e}		1.00e+06		#指数记法 保留小数点后两位
13			{:>10d}		        13		#右对齐 (默认, 宽度为10)
13			{:<10d}		13        		#左对齐 (宽度为10)
13			{:^10d}		    13    		#中间对齐 (宽度为10)
11 			'{:b}'			1011 			#进制转换
11 			'{:d}'			11 				#十进制
11 			'{:o}'			13 				#八进制
11 			'{:x}'			b 					#十六进制
11 			'{:#x}'		0xb				#另一种十六进制
11 			'{:#X}'		0XB				#另一种十六进制大写
runoob	"{}{{}}"		runoob{}		#输出{和}

print("*%format" %(*var))#% 格式运算符
%s #格式化字符串
%d #格式化整数
%f #格式化浮点数字 可指定小数点后的精度
%e #用科学计数法格式化浮点数
%c #格式化字符及其ASCII码
%u #格式化无符号整型
%o #格式化无符号八进制数
%x #格式化无符号十六进制数
%X #格式化无符号十六进制数 大写
%E #作用同%e 用科学计数法格式化浮点数 大写
%g #%f和%e的简写
%G #%f 和 %E 的简写
%p #用十六进制数格式化变量的地址
* #定义宽度或者小数点精度
- #用做左对齐
+ #在正数前面显示加号
<sp>#在正数前面显示空格
#	#在八进制数前面显示零('0') 在十六进制前面显示'0x'或者'0X'(取决于用的是'x'还是'X')
0	#显示的数字前面填充'0'而不是默认的空格
%	'%%'输出一个单一的'%'
%y #两位数的年份表示（00-99）
%Y #四位数的年份表示（000-9999）
%m #月份（01-12）
%d #月内中的一天（0-31）
%H #24小时制小时数（0-23）
%I #12小时制小时数（01-12）
%M #分钟数（00=59）
%S #秒（00-59）
%a #本地简化星期名称
%A #本地完整星期名称
%b #本地简化的月份名称
%B #本地完整的月份名称
%c #本地相应的日期表示和时间表示
%j #年内的一天（001-366）
%p #本地A.M.或P.M.的等价符
%U #一年中的星期数（00-53）星期天为星期的开始
%w #星期（0-6） 星期天为星期的开始
%W #一年中的星期数（00-53）星期一为星期的开始
%x #本地相应的日期表示
%X #本地相应的时间表示
%Z #当前时区的名称
(var)	#映射变量(字典参数)
m.n.	#m 是显示的最小总宽度,n 是小数点后的位数(如果可用的话)
print("nHex=%x,nDec=%d,nOct=%o" %(0xFF,0xFF,0xFF))
>>>nHex=ff,nDec=255,nOct=377
print('%10.3f' % math.pi) #字段宽10 四舍五入三位小数3 默认右对齐
>>>     3.142
print('%010.3f' % math.pi) #用0填充空白 字段宽10 四舍五入三位小数3
>>>000003.142
print('%-10.3f' % math.pi) #左对齐
>>>3.142
print('%+f' % math.pi) #显示正负号
>>>+3.141593

#string
import string
string.ascii_letters#获取所有ascii码中字母字符的字符串（包含大写和小写）
string.ascii_uppercase#获取所有ascii码中的大写英文字母
string.ascii_lowercase#获取所有ascii码中的小写英文字母
string.digits#获取所有的10进制数字字符
string.octdigits#获取所有的8进制数字字符
string.hexdigits#获取所有16进制的数字字符
string.printable#获取所有可以打印的字符
string.whitespace#获取所有空白字符
string.punctuation#获取所有的标点符号

#ASCII 美国标准信息交换代码 定制了128个常用字符 主要是英文 数字 标点符号及键盘中其他按键对应的整数值
#A~Z	65~90
#a~z	97~122
#0~9	48~57
str=chr(index)#转换字符索引为字符 自动识别属于哪种编码
ascii=ord(str)#转换字符为字符索引 自动识别属于哪种编码

#str
trans_dict=str.maketrans(orgin_strs,sub_strs)#制作翻译表 将 orgin_strs 中的每一个字符 替换成 sub_strs 中对应索引的字符 长度必须相同 返回一个ASCII或utf-8字符索引替换前后一一对应的字典
transed_str=str.translate(trans_dict)#用trans_dict替换str的所有字符
str=repr(var_name)#返回"var_name"
print(repr(s)) >>> '1 2\t 3\n 4'
class=eval(class_str)#计算字符中的表达式 或 返回字符表示的类
c=eval('pow(2,3)')#pow(x,y)返回x^y
>>>c=8
a=eval("{1:'xx',2:'yy'}")
>>> a={1:'xx',2:'yy'}
exec(str_expression)#执行字符串中的语句
execfile(filename)#执行filename文件
str.lower()
str.upper()
str.title()
str.capitalize()#only title the first word
str.swapcase()#swap lower and upper
str.center(width,str_used)#str居左居右居中 用str_used填充
str.ljust(width,str_used)
str.rjust(width,str_used)
numstr.zfill(width,default=0)#填充0占位

str.isdecimal()#判断给定字符串是否全为数字
str.isalpha()#判断给定的字符串是否全为字母
str.isalnum()#判断给定的字符串是否只含有数字与字母
str.isupper()#判断给定的字符串是否全为大写
str.islower()#判断给定的字符串是否全为小写
str.istitle()#判断给定的字符串是否符合title()
str.isspace()#判断给定的字符串是否为空白符（空格 换行 制表符）
str.isprintable()#判断给定的字符串是否为可打印字符（只有空格可以 换行 制表符都不可以）
str.isidentifier()#判断给定的字符串是否符合命名规则（只能是字母或下划线开头 不能包含除数字 字母和下划线以外的任意字符

str.startswitch(str)#字符以str开头 返回True False
str.endswitch(str)#字符以str结尾 返回True False 可用于判断扩展名
str.count(str,start,end)#计数str在字符串中出现的次数
l,c,r=str.partition(str)#由字符串中str位置切分成三个字符串
str=str.join(iter)
list=str.split(split_str)#返回分割后的列表
str=str.strip(default_str=' ')#删除字符串头尾指定的所有str 不能去除所有
str=str.replace(old,new,times)#去除所有
str.encode(encoding)#编码
str.decode(encoding)#解码
#encoding
encoding='UTF-8'#一个汉字占3个字节
encoding=ASCII#一个字符占1个字节
encoding=ANSI
encoding=GBK#一个汉字占2个字节
encoding=GB2312
encoding=GB18030
encoding=UNICODE
'\n'#换行
'\t'#tab
'\r'#回车 将光标移动到本行开头
#enter=\r\n

#* 和 ** 有四类用法
*  #代表乘法
** #代表乘方

*args **kwargs #主要用于函数定义 都是 python 中的可变参数
def func(arg,arg=default,*args, **kwargs):
#如果同时使用 *args 和 **kwargs 时 必须 *args 参数列要在 **kwargs 之前
*args #表示任何多个无名参数 它本质是一个 tuple
def func(name, *args):
	print('你好:', name)
	for i in args:
		print("你的宠物有:", i)
>>> func("Geek", "dog", "cat")
你好: Geek
你的宠物有: dog
你的宠物有: cat

**kwargs #表示关键字参数 它本质上是一个 dict
def func(**kwargs):
	for key, value in kwargs.items():
		print("{0} 喜欢 {1}".format(key, value))
>>> func(Geek="cat", cat="box")
Geek 喜欢 cat
cat 喜欢 box

#用 *args 和 **kwargs 调用函数 类似对元组和字典进行解引用
def func(data1, data2, data3):
	print(data1, data2, data3)
args=("one", 2, 3)
kwargs={"data3": "one", "data2": 2, "data1": 3}
func(*args) >>> one  2 3
func(**kwargs) >>> one  2 3

*#序列解包 **不能
a, *b,c= 0, 1, 2, 3,4
>>>a=0
>>> b=[1,2, 3]
>>> c=4
a,*b=1
>>>a=1,b=[]

s='ABCDEFGH'
while s:
	x, *s=s
	print(x, s)
>>>
A ['B', 'C', 'D', 'E', 'F', 'G', 'H']
B ['C', 'D', 'E', 'F', 'G', 'H']
C ['D', 'E', 'F', 'G', 'H']
D ['E', 'F', 'G', 'H']
E ['F', 'G', 'H']
F ['G', 'H']
G ['H']
H []

#随机
import random
random.seed(num)# 初始化给定的随机数种子 默认为当前系统时间 随机出一个数字 且保证以后的随机数都是这个数字
random.random()#生成一个[0.0,1.0)之间的随机小数
random.randint(a,b)#int [a,b]
random.randrange(a,b,k)#int [a,b) step k
random.getrandibits(k)#int  生成一个k比特长的随机整数
random.uniform(a,b)#float [a,b]
random.choice(iter)#choice one from iter
random.shuffle(iter)#打乱iter

#排列组合
import itertools
#repeat 表示组合位数
i_list=itertools.product(iter, digit)#笛卡尔积
i_list=itertools.permutations(iter, digit)#排列 digit是排列组合的位数 若不填 默认返回全部排列组合情况
i_list=itertools.combinations(iter, digit)#组合
i_list=itertools.combinations_with_replacement(iter, digit)#组合 含自身重复

#连接词
and #有一个False 输出就为False
or #有一个True 输出就为True
if (var1 in sth) or (var2 in sth):#判断两个变量和另一个变量之间的关系
不要使用这种语句 判断会出错#if (var or var) in sth:

#列表全为False
any(iter)#iter中至少一个值为True 就返回True
#列表全为True
all(iter)#iter中每个值为True 才返回True

#程序中断
break #for or while 跳出并结束一层循环
continue #for or while 跳回到该层循环开头
sys.exit() #结束程序 import sys
return #def 返回值并跳出函数
yield #def 返回值并跳出函数 下次直接从yield后的一条语句开始运行函数
pass #* 跳过该处代码

#运算符
** #^
% #取余数
// #取整数商数
** % // / *-+#数学运算符优先级
+= -= *= /= %=#符号运算简写
== = #比较值
is #比较内存地址

#运算函数
complex(real,imagnum)#创建复数(real + imagnumj)
complex('realnum+imagnumj')#将字符转为复数
quot,rem=divmod(7,2) #返回商和余数
r=round(float,digit)#返回digit位小数的float 四舍六入五成双
from decimal import Decimal, ROUND_HALF_UP#实现精确的四舍五入
answer_num = Decimal('11.245').quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
>>>11.25
answer_num = Decimal(str(float)).quantize(Decimal('0.digit'), rounding=ROUND_HALF_UP)
r=format(float,'.0%')#返回0位小数的百分数
abs_value=abs(num)#返回绝对值
r=pow(x,y,z=None)#返回x的y次方 如果z存在 返回pow(x,y)%z
r=sum(iter,value=0)#返回iter相加后再和value相加的结果
cmp(x,y)#如果 x < y 返回 -1, 如果 x == y 返回 0, 如果 x > y 返回 1
binary=bin(int)#返回int二进制
hash(obj)#返回哈希值
hex()#十进制转为十六进制 字符串形式表示
oct()#十进制转为八进制 字符串形式表示
id(obj)#返回对象内存地址
dir(obj)#返回对象属性 函数列表

#运算模块
import math#数学函数
import cmath#复数函数

#迭代器 列表 字符串等有顺序的都是迭代器
iter(iter)#生成迭代器
next(iter,default)#返回迭代器的下一个项目 next() 函数要和生成迭代器的 iter() 函数一起使用
s=slice(start,stop,step)#切片
iter[s]#截取片段
content_iter=(line for line in open(file))#返回迭代器 而不是file的内容 只能用()
print(next(content_iter))#每次返回下一行line
print(next(content_iter))

#布尔值
bool(var)#返回布尔值
0 0.0 None [] () {} '' "" return None#布尔值为 False
assert True_code,False_code#断言 程序正常运行调用True_code,程序出错调用False_code

#等效
var='str'
if var == '':
if not var != '':
if var == False:
if not var:
if False:
#等效
var=1
while var != 0:
while var != False:
while var == True:
while var:
while True:

#范围
range(0,3) # 0 1 2 不含3
range(3,0,-1) # 3 2 1

#全局变量
global#全局作用域不能使用局部作用域变量
#local #局部作用域能使用全局作用域变量
def func():
	global var#声明全局变量
	var #在函数内部创建全局变量
locals()#以字典类型返回当前位置全部局部变量
globals()#函数会以字典类型返回当前位置的全部全局变量
def func():
	def inner():
		nonlocal var#声明非局部变量
		var

#异常处理
try: #可能异常的代码
except ERROR as e: #处理异常 捕捉异常
else:#没有发生异常时的代码
finally:#是否发生异常都执行的代码
#写代码时要主动考虑使用程序时会产生的异常情况
raise Exception()#引起异常
raise ValueError(str)#通用自定义异常提示
class Cust_Exc(Exception):#自定义异常
	"""docstring for Cust_Exc"""
	def __init__(self, msg):
		self.msg=str
	def __str__(self):
		return self.msg
raise Cust_Exc(str)

#比较大小
max(iter)
min(iter)

#列表
for index,element in enumerate(list,num):#索引和元素 num可以不写 num表示index的初始值 默认index=num=0 用enumerate代替range
	list[index]=new_element#修改元素的方法一 不能用for循环直接更改
for index in range(len(list)):#修改元素的方法二 不能用for循环直接更改 不建议用这种
	list[index]=new_element
l=list()
list1_2=list1+list2
list1_2=list(set(list1)-set(list2))#list相减
a=[1,2,3,4] a[0:-1]=[1,2,3] a[1:4]=[2,3,4] a[:2]=[1,2]
list[*start:end:step,]#切片含头不含尾 多层之间用,分隔
list=[[*ele],[*ele],[*ele],[*ele],[*ele],[*ele]]
list[1:3,:3]#二维列表 list中索引为1-2的元素中 索引为0-2的元素
list[::-1]#反转字符串
del list[index]
list.index(value) #返回value第一次出现的索引
list.append(value) #value will be the last in the list
list.extend(iter) #iter will be the last in the list
list.insert(index,value) #将指定对象插入列表的指定位置
list.remove(value) #value firstly appearing will be removed from the list
poped=list.pop(index)
poped=list.pop()#default index -1
list.sort()#the values of list will be arranged in order (must be the same type)
list.sort(reverse=True) #the values of list will be arranged in order in reverse
list.sort(key=func)#func need to definite
list.sort(key=str.lower) #the values of list will be arranged in order (ignore a or A type)95
def take_sec(itr):
	return itr[1]
[(2, 2), (3, 4), (4, 1), (1, 3)].sort(key=take_sec)#指定第二个元素排序
#列表中的值可改 元组和字符不可
#不可变的值(字符元组整型etc) python变量保存本身 可变的值(列表字典) 引用本身
list=[x func_condition]#列表条件 处理元素
list=[x**x for x in range(n)]
matrix=[[1,2,3],[4,5,6]]
flat=[x for row in matrix for x in row]#二维转一维
squared=[[x**2 for x in row]for row in matrix]#对每个元素平方

#排序
new=sorted(iter,key=key_func)
new=reversed(iter)
class Student:
	def __init__(self, name, grade, age):
		self.name=name
		self.grade=grade
		self.age=age
	def __repr__(self):
		return repr((self.name, self.grade, self.age))
students_list=[Student('john', 'A', 15),
Student('jane', 'B', 12),Student('dave', 'B', 10),]
s=sorted(students_list, key=lambda student: student.age)   # sort by age
>>>[('dave', 'B', 10), ('jane', 'B', 12), ('john', 'A', 15)]
from operator import attrgetter
s=sorted(students_list, key=attrgetter('grade', 'age'))#先以grade再以age来排序
>>>[('john', 'A', 15), ('dave', 'B', 10), ('jane', 'B', 12)]

#字典
a=dict()
dict[key]=value
reversed_dict={value:key for key,value in dict.items()}#反转字典
key in dict #CANNOT use value in dict
value in dict.values()
key in dict.keys()
key-value in dict.items()
value=dict.get(key,defaultvalue=None) #如果键不存在于字典中 将会返回默认值
del dict[key]
dict.clear()
dict.update(dict2)
poped=pop(key)
poped=popitem()#return and del the last item
d=dict(zip(iter1,iter2))
dict.setdefault(key,defaultvalue)#如果键不存在于字典中 将会添加键并将值设为默认值
for key,value in [(key,value),]:#元组列表转字典形式
	dict[key]=value

#元组 unmodifiable
t=tuple()
(0,1,200000)<(0,3,4) #return True 有一个不同且能比较 比较后返回结果 不再继续比较后续元素
func(*tuple) #拆分元组

#集合 可用于去重复元素 但会打乱顺序
{}=set()
list1_2=list(set(list1)-set(list2))#list相减
a=frozenset(iter)#返回一个不可变的集合

#函数
def func(*args):#任意多个参数
def func(a,ar,arg=default):#defaultvalue
def func(arg,arg=default,*args, ):#参数 参数默认值 可变参数(元组列表类型)
def func(**kwargs):#可变参数(字典类型)
	print(a,b,c)#输出字典值1 2 3
func(**{'a': 1, 'b': 2, 'c': 3})#使用**时 字典键必须和函数参数名字一致 否则出错
map(func,iterable)#函数映射到迭代器
list=filter(func,iter)#过滤出符合func的元素

#压缩迭代器 一一对应 取最短的iter
z=zip(*iter)#z=(val1_1,val2_1)(val1_2,val2_2)(val1_3,val2_3)
z=list(zip(iter1),iter2)#[(1,1),(2,2)]
z=itertools.zip_longest(*iter)#取最长的iter

#时间
unix时间戳 从1970.01.01 0:0:0 开始到现在一共的秒数 只能表示1970-2038的时间
1 分钟 	60
1 小时 	3600
1 天 		86400
1 周 		604800
1 月 (30.44 天) 	2629743
1 年 (365.24 天) 	31556736
import time
time.time() #返回当前unix时间戳(float) time模块中的函数默认值都为time.time()
time.sleep(sec)#程序推迟执行的秒数
struct_time_tuple=time.gmtime(unix_time_stamp)#unix时间戳转0时区struct_time类
时间戳(float)转字符串(str) 转出来的时间是UTC时区（0时区）的struct_time 但是我们计算机显示的是东八区时间（+8） 所以的得到的struct_time+8即为现在计算机显示的时间
时间转换与所在时区有关
print(struct_time_tuple)
>>>time.struct_time(tm_year=2016, tm_mon=4, tm_mday=7, tm_hour=2, tm_min=55, tm_sec=45, tm_wday=3, tm_yday=98, tm_isdst=0)
0 tm_year #年份 其值等于实际年份减去1900
1 tm_mon #月份（从一月开始 0代表一月）-取值区间为[0,11]
2 tm_mday #一个月中的日期-取值区间为[1,31]
3 tm_hour #时-取值区间为[0,23]
4 tm_min #分-取值区间为[0,59]
5 tm_sec #秒 – 取值区间为[0,59]
6 tm_wday #星期 – 取值区间为[0,6] 其中0代表星期一 1代表星期二 以此类推
7 tm_yday #从每年的1月1日开始的天数 – 取值区间为[0,365] 其中0代表1月1日 1代表1月2日 以此类推
8 tm_isdst #夏令时标识符 实行夏令时的时候 tm_isdst为1 不实行夏令时的时候 tm_isdst为0 不了解情况时 tm_isdst()为-1
local_struct_time_tuple=time.localtime(unix_time_stamp)#unix时间戳转本地struct_time类
secs=time.mktime(local_struct_time_tuple)#本地struct_time类转unix时间戳 mktime 输入的日期是带时区的 返回的值才是不带时区的 直接用 gmtime(0) 生成的 date 组装成数组输入 mktime 则会报超出范围的错误
str_time=time.strftime(str_format,struct_time_tuple)#struct_time转字符串时间格式
struct_time_tuple=time.strptime(str_time,str_format)#字符串时间格式转struct_time
'Tue Feb 17 09:42:58 2009'=time.asctime(time.gmtime() or time.localtime())#接受时间元组并返回一个可读的形式
'Tue Feb 17 10:00:18 2013'=time.ctime(sec)#相当于time.asctime(localtime(secs))
1530000078.0=time.mktime(time.strptime('2018-06-26 16:01:18', '%Y-%m-%d %H:%M:%S'))#字符串时间格式转为unix时间戳
datetime_class=datetime.strptime('2018-06-26 16:01:18', '%Y-%m-%d %H:%M:%S')#字符串转为datetime类(datetime.datetime)
'Sat Mar 28 22:24:24 2016'=time.strftime('%a %b %d %H:%M:%S %Y', time.gmtime(1530000078.0))#unix时间戳转为字符串时间格式
datetime_class=datetime.strptime(time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(1530000078.0)), '%Y-%m-%d %H:%M:%S')#unix时间戳转datetime类
'2018-06-26 16:01:18'=time.strftime("%Y-%m-%d %H:%M:%S",time.gmtime(1530000078.0))#datetime类转为字符串时间格式
time_stamp=float(time.mktime(time.strptime(time.strftime("%Y-%m-%d %H:%M:%S",time.gmtime(1530000078.0)), '%Y-%m-%d %H:%M:%S')))#datetime类转为unix时间戳
str_time=time.strftime("%Y-%m-%d %H:%M:%S",datetime(year=1900, month=1, day=1)+datetime.timedelta(days=excel_datetime)) #datetime类转为字符串时间格式 excel_datetime 从excel读出的float日期 eg 44152.0146412037
#常用时间处理
today=datetime.date.today()#今天
yesterday=today-datetime.timedelta(days=1)#昨天
last_month=today.month-1 if today.month-1 else 12#上个月
time_stamp=time.time()#当前时间戳
datetime.datetime.fromtimestamp(time_stamp)#时间戳转datetime
int(time.mktime(today.timetuple()))#datetime转时间戳
today_str=today.strftime("%Y-%m-%d")#datetime转字符串
today=datetime.datetime.strptime(today_str, "%Y-%m-%d")#字符串转datetime
today + datetime.timedelta(hours=8)#补时差

#datetime
import datetime
datetime.date.today()#返回当前字符串日期
datetime.date.fromtimestamp(time_stamp)#时间戳转字符串日期
d=datetime.date(year,month,day)
d.year
d.month
d.day
d.replace(newyear,newmonth,newday)
d.timetuple()#返回struct_time
d.isocalendar()#返回格式如(year month day)的元组
d.isoformat()#返回格式如'YYYY-MM-DD’的字符串
t=datetime.time(hour,minute,second)
t.hour
t.minute
t.second
t.replace(newhour,newminute,newsecond)
t.isoformat()#返回格式如'HH:MM:SS’的字符串
datetime.datetime.today()#返回一个表示当前本地时间的datetime对象
datetime.datetime.now(time_zone)#返回一个表示当前本地时间的datetime对象 如果提供了参数time_zone 则获取tz参数所指时区的本地时间
datetime.datetime.utcnow()#返回一个当前utc时间的datetime对象#格林威治时间
datetime.datetime.fromtimestamp(time_stamp,time_zone)#根据时间戳创建一个datetime对象 参数tz指定时区信息
datetime.datetime.utcfromtimestamp(time_stamp)#根据时间戳创建一个datetime对象
datetime.datetime.combine(date,time)#根据date和time 创建一个datetime对象
datetime.datetime.strptime(date_string,format)#将格式字符串转换为datetime对象
dt=datetime.datetime(year,month,day,hour,minute,second)
dt.year month day hour minute second microsecond timezone
dt.date()#获取date对象
dt.time()#获取time对象
dt.replace(year, month, day, hour, minute, second, microsecond, timezone)
dt.timetuple()
dt.utctimetuple()
dt.isocalendar()
dt.isoformat(sep)
dt.ctime()#返回一个日期时间的C格式字符串 等效于time.ctime(time.mktime(dt.timetuple()))
dt.strftime(format)
dt.number_format#'yyyy-mm-dd h:mm:ss'
datetime.timedelta(days=v,seconds=v,weeks=v,minutes=v,hours=v,microseconds=v)#时间计算 不能计算年月 只能和datetime类计算

#dateutil 用于datetime时间处理
from dateutil.parser import parse
parse('date_string')#任意正常日期格式转为datetime.datetime(year,month,day,hour,minute,sec)
#年份放在前面时 只能按年-月-日的顺序
#分隔符为逗号时 只有月-日时 要把月放在前面 有年份时 年份要放在后面
parse("2018/10/21")
>>> datetime.datetime(2018, 10, 21, 0, 0)
from dateutil.rrule import rrule
rrule(freq, dtstart=None, interval=1, wkst=None,count=None, until=None, bysetpos=None,bymonth=None, bymonthday=None, byyearday=None, byeaster=None,byweekno=None, byweekday=None, byhour=None, byminute=None, bysecond=None)
freq#单位 可以是 YEARLY, MONTHLY, WEEKLY,DAILY, HOURLY, MINUTELY, SECONDLY 即年月日周时分秒
dtstart,until#是开始和结束时间
wkst#周开始时间
interval#间隔
count#只保留几个
byxxx#指定匹配的周期 比如byweekday=(MO,TU)则只有周一周二的匹配 byweekday可以指定MO,TU,WE,TH,FR,SA,SU 即周一到周日
list(rrule(DAILY,byweekday=(SA,SU),dtstart=parse('2018-11-1'),until=parse('2018-11-5')))#只要周六和周日的
>>>[datetime.datetime(2018, 11, 3, 0, 0),
datetime.datetime(2018, 11, 4, 0, 0)]
rrule(freq,dtstart=parse('date'),until=datetime.date.today()).count()#计算某个日期到今天相差freq数

#calendar 日历
import calendar
calendar.calendar(year,w=2,l=1,c=6)#返回一个年历 3个月一行 间隔距离为c  每日宽度间隔为w字符 l是每星期行数 不填会设置为默认间距
calendar.isleap(year)#是闰年返回 True 否则为 False
calendar.leapdays(y1,y2)#返回在Y1 Y2两年之间的闰年总数
calendar.month(year,month,w=2,l=1)#返回一个多行字符串格式的year年month月日历 两行标题 一周一行 每日宽度间隔为w字符 l是每星期的行数 不填会设置为默认间距
calendar.timegm(tupletime)
#和time.gmtime相反接受一个时间元组形式 返回该时刻的时间戳（1970纪元后经过的浮点秒数）类似time.mktime()
calendar.monthrange(year,month)#返回(星期几,这月有几天)
days_list=calendar.mdays#当年每个月的天数组成的列表
#判断类型
type(obj)#return type
isinstance(n,type)#return True or False

#复制 拷贝
import copy
copy.copy()#赋值 拷贝
copy.deepcopy()#深拷贝

origin=[1, 2, 3, 4, ['a', 'b']] #原始对象
assigned=origin#赋值 传对象的引用 两者的子对象(列表里的每个元素)父对象(列表本身)是同一个 让两个变量的数据相同 且是同一个变量
copied=copy.copy(origin)#对象拷贝 浅拷贝 两者的子对象是同一个 父对象是独立的两个 让两个变量的数据相同 是同一个数据 但不是同一个变量
deepcopied=copy.deepcopy(origin)#对象拷贝 深拷贝 两者的子对象父对象分别是值相同的独立两个 让两个变量的数据在最初相同 但不是同一个数据 也不是同一个变量
origin.append(5)#修改对象origin
origin[4].append('c') #修改对象origin中的['a', 'b']数组对象
>>>origin=[1, 2, 3, 4, ['a', 'b', 'c'], 5]
>>>assigned=[1, 2, 3, 4, ['a', 'b', 'c'], 5]
>>>copied=[1, 2, 3, 4, ['a', 'b', 'c']]
>>>deepcopied=[1, 2, 3, 4, ['a', 'b']]

#代码优化 略
import profile

from logging import basicConfig,DEBUG,debug,CRITICAL,disable#日志模块 记录调试信息 可以直接写在文件头
#disable(CRITICAL)#禁用日志 在程序完成之后 将最前面的#取消
basicConfig(level=DEBUG, format='%(levelname)s: %(message)s. [%(lineno)d]%(filename)s <%(asctime)s>',filename='debug.log',filemode='w')#配置日志输出格式 以'w'模式写入debug.log储存在程序同级目录
debug(str)#用来记录调试信息 str将表示在%(message)s中

#format
%(levelno)s#打印日志级别的数值
%(levelname)s#打印日志级别名称
%(pathname)s#打印当前执行程序的路径 等于sys.argv[0]
%(filename)s#打印当前执行程序名
%(funcName)s#打印日志的当前函数
%(lineno)d#打印日志的当前行号
%(asctime)s#打印日志的时间
%(thread)d#打印线程ID
%(threadName)s#打印线程名称
%(process)d#打印进程ID
%(message)s#打印日志信息

#日志级别 				日志函数
logging.DEBUG 		logging.debug()#小细节
logging.INFO 			logging.info()#一般信息
logging.WARNING	logging.warning()#警告
logging.ERROR 		logging.error()#错误
logging.CRITICAL	logging.critical()#致命错误

#另一种Bug处理方式 idle调试器debugger 略

#读写文件
f=open(filename,mode,encoding='UTF-8')#以mode模式 以UTF-8编码 打开或创建filename 不用utf-8可能引起异常
#encoding参数 Window上默认字符编码为GBK Linux上默认字符编码为UTF-8
encoding='UTF-8'#一个汉字占3个字节
encoding=ASCII#一个字符占1个字节
encoding=ANSI
encoding=GBK#一个汉字占2个字节
encoding=GB2312
encoding=GB18030
encoding=UNICODE
str.decode(encoding)#若字符乱码 修改解码方式
#mode
'r'#以 只读模式 打开文件 并将文件指针指向文件头 如果文件不存在会报错
'w'#以 只写模式 打开文件 并将文件指针指向文件头 如果文件存在则将其内容清空 如果文件不存在则创建
'a'#以 只追加可写模式 打开文件 并将文件指针指向文件尾部 如果文件不存在则创建
'r+'#在r的基础上增加了可写功能,会覆盖当前文件指针所在位置的字符 如原来文件内容是"Hello World" 打开文件后写入"hi"则文件内容会变成"hillo, World"
'w+'#在w的基础上增加了可读功能,在打开文件时就会先将文件内容清空
'a+'#在a的基础上增加了可读功能, 并将文件指针指向文件尾 只能写到文件末尾（无论当前文件指针在哪里）
如果以此种方式打开直接读取文档 会返回空值 使用f.seek(0, 0) 使文件指针指向文档开头
'b'#读写二进制文件（默认是t 表示文本） 需要与上面几种模式搭配使用 如ab wb, ab, ab+（POSIX系统 包括Linux都会忽略该字符）
str_read=f.read()#读取内容
f.close()#关闭文件
with open(filename,mode,encoding) as f:#f.open()和f.close()的简化
	for line in f:#推荐这种方式 内存消耗小
		print(line,end='')#避免print方法造成的换行
content=f.read()#一次读取文件所有内容 返回一个str
content=f.read(int_size)#每次最多读取指定长度的内容 返回一个str 在Python2中size指定的是字节长度 在Python3中size指定的是字符长度
content_list=f.readlines()#一次读取文件所有内容 按行返回一个list
line_string=f.readline()#每次只读取一行内容
f.seek(n)#将文件指针移动到指定字节的位置
f.seek(0, 0)#将文件指针移动到文档开头位置
f.tell()	#获取当前文件指针所在字节位置
str_amount=f.write(str)#返回写入字符的数量并写入字符
next_line=f.next()#返回文件下一行 这个方法也是file对象实例可以被当做迭代器使用的原因
f.writelines(iter)#向文件写入一个字符串列表 字符串需要自己添加换行符

#shelve 二进制储存数据
import shelve
shelf_file = shelve.open('shelf_name')#打开或创建shelf_name
shelf_file['data_name'] = data#shelf_file可以看作一个字典
shelf_file.close()#会在程序同居目录生成.bak .dat .dir三个文件进行二进制储存数据

#操作系统
#os operating system
import os
.#当前工作目录简写 .\
..#当前工作目录的上一级目录简写 ..\
os.system(r'str')#在cmd运行str语句
str=os.popen(r'str')#在cmd运行str语句 且返回str
cwd=os.getcwd()#返回当前工作目录
cwd=os.chdir(path)#改变当前工作目录
fullpath=os.path.abspath(path)#返回目录绝对路径
print(os.path.exists(path))#返回目录存在布尔值
print(os.path.isabs(path))#返回目录存在布尔值
if os.path.exists(out_dir)==False: os.mkdir(out_dir)#如果输出文件夹不存在 创建out_dir文件夹
os.mkdir(dirname)#创建一层文件夹
os.makedirs('D:\\folder1\\folder2\\folder3\\...')#创建多层文件夹
os.remove(fullpath)#只用于移除文件和文件夹
os.rename(old_file_path, new_file_path)#只能对相应的文件进行重命名, 不能重命名文件的上级目录名
os.renames(old_file_path, new_file_path)#是os.rename的升级版, 既可以重命名文件, 也可以重命名文件的上级目录名 不是新建目录
os.unlink(fullpath)#永久删除 path 处的文件
os.rmdir(fullpath)#永久删除 path 处的文件 只能删除空文件夹
os.removdirs(fullpath)#只能永久删除空文件夹 目录 将整个目录删除
os.utime(fullpath, (accessed_time_stamp,modified_time_stamp))#修改文件的访问时间 改动时间 时间格式为unix时间戳
print(os.path.isdir(fullpath))#返回是否目录或文件夹布尔值
print(os.path.isfile(fullpath))#返回是否文件布尔值
files_list=os.listdir(dirname)#返回该目录内的第一层文件夹和文件列表
import fnmatch
for filename in os.listdir(dirname):
	if fnmatch.fnmatch(filename,'*.txt'):pass#匹配目录中所有.txt文件
(root,dirs_list,files_list)=os.walk(root,topdown=True)#返回该目录内的所有文件名列表 包含所有级目录的元组
#topdown 默认为True 会优先遍历root目录 False 优先遍历root子目录
#root是文件本身目录 相当于dirname
#dirs_list是该文件夹所有目录名字的列表(不含子目录)
#files_list是该文件夹所有文件名字的列表(不含子目录)
for root, dirs, files in os.walk(root):
	for name in files:#拼接文件名
		print(os.path.join(root, name))
	for name in dirs:#拼接目录名
		print(os.path.join(root, name))
fullpath=os.path.join(dirname,filename)#拼接目录和文件名 返回绝对路径
(dirname,filename)=os.path.split(fullpath)#返回路径与文件名组成的元组 不含/
(file_name,extension)=os.path.splitext(filename)#返回文件名字和后缀(扩展名 含.)组成的元组
['C:', 'Windows', 'System32', 'calc.exe']='C:\\Windows\\System32\\calc.exe'.split(os.path.sep)#根据使用的系统判断使用\还是/分割
dirname=os.path.dirname(fullpath)#返回目录路径 D:\folder1\folder2
filename=os.path.basename(fullpath)#返回文件名 *.*
accessed_time=os.path.getatime(filename)#输出最近访问时间 累计秒格式
created_time=os.path.getctime(filename)#输出文件创建时间 累计秒格式
modified_time=os.path.getmtime(filename)#输出最近修改时间 累计秒格式
clear_time=time.gmtime(os.path.getmtime(filename))# 以struct_time形式输出最近修改时间
>>>clear_time=time.struct_time(tm_year=2016, tm_mon=4, tm_mday=7, tm_hour=2, tm_min=55, tm_sec=45, tm_wday=3, tm_yday=98, tm_isdst=0)
file_size=os.path.getsize(filename)#输出文件大小 以字节为单位

from pyperclip import copy,paste#访问剪贴板
copy(str)#复制到剪切板 只支持文本内容
str=paste()#粘贴文本

import sys
script_name=sys.argv[0]#0代表程序本身完整目录 1: 代表从cmd接受的每个参数 以空格分割
user_argv=sys.argv[1]
user_argv_list=sys.argv[1:]
cmd输入 script.py -argv1 a2 d
>>>argv_list=['script.py','-argv1','a2','d']

import shutil#复制 移动 改名 删除文件
destination_file_path=shutil.copy(source_file_path,destination_file_path)#将 source_file_path 复制到 destination_file_path 返回 destination_file_path 只能复制文件
destination_path=shutil.copytree(source_path,destination_path)#将 source_path 复制到 destination_path 返回 destination_path 用于复制文件夹 包括文件夹内的所有文件和文件夹
destination_path=shutil.move(source_path,destination_path)#将 source_path 移动到 destination_path 返回 destination_path 会覆盖和 destination_path 同名的文件 destination_path 中的文件夹目录必须存在 否则会异常 如果文件名不存在 会创建
shutil.rmtree(path)#永久删除path文件夹 包含所有子文件和文件夹

import send2trash#需要安装模块
send2trash.send2trash(filename)#移动到回收站

#fnmatch 主要用于文件匹配 目录匹配 正则表达的简化模块
import fnmatch,os
* #匹配任意字符
? #匹配单个字符
[seq] #匹配指定范围内字符
[!seq] #匹配不在指定范围内字符
fnmatch.fnmatch(filename, 'pattern')#匹配符合pattern的filename
fnmatch.fnmatchcase(filename, 'pattern')#强制区分大小写进行匹配 和fnmatch一样
fnmatch.filter(list, 'pattern')#在list中匹配符合pattern模式的元素
regex=fnmatch.translate('pattern')#将'pattern'转为正则表达式进行匹配
for filename in os.listdir(dirname):
	if fnmatch.fnmatch(filename,'*.txt'):pass#匹配目录中所有.txt文件

#glob 主要用于文件匹配
import glob
*#匹配0个或多个字符
?#匹配单个字符
[]#匹配指定范围内的字符 如 [0-9]匹配数字
for n in glob.glob('pattern'):#同时获取所有的匹配路径
for n in glob.iglob('pattern'):#一次只获取一个匹配路径

import zipfile#压缩模块
zip_file=zipfile.ZipFile('zip_file.zip')#创建 ZipFile 对象
zip_dir_list=zip_file.namelist()#返回zip文件里的所有文件目录组成的列表 如
zip_file.zip
	cats_folder
		catnames.txt
		zophie.jpg
	spam.txt
>>>zip_dir_list=['spam.txt', 'cats_folder/', 'cats_folder/catnames.txt', 'cats_folder/zophie.jpg']
spam_info=zip_file.getinfo('spam.txt')#返回spam.txt的信息对象
spam_info.file_size#原文件大小 字节单位
spam_info.compress_size#压缩后文件大小 字节单位
zip_file.extractall('path')#解压到path目录 不存在会被创建 默认为程序同级目录
outpath=zip_file.extract('filename','path')#解压单个filename到path目录 不存在会被创建 默认为程序同级目录 返回path目录
zip_file.write('add_filename',compress_type=zipfile.ZIP_DEFLATED)#用deflate算法将add_filename压缩进zip_file
zip_file.close()#压缩包不使用一定要关闭

#类 任何对象都可以定义类 例如 学校-年级-班级-学生-成绩 每一层都是一个类 数据层次太多必须用类 否则数据难以处理
class ClassName(object):
	"""docstring for ClassName"""
	def __init__(self, arg=defaultvalue):
		super(ClassName, self).__init__()
		self.arg=arg
		self.__private_arg=0  #私有变量 不能在类的外部调用 只能在类本身调用 用__在变量开头声明
		self._protected_arg=0 #保护变量 只能本身和子类使用 不能作为模块引入 不能使用import访问 在变量开头声明
		self.public_arg=0# 公开变量
	def 	__special_func__(self):#特殊方法 系统定义 用__在函数开头结尾声明
	def count(self):
		self.__private_arg += 1# 私有变量
	def __str__(self):#当使用print(ClassName)时将调用该方法
		return '{}'.format(self.arg)
	def __del__(self):#销毁对象
		class_name=self.__class__.__name__
		print('has been deled')
	def __customize_func__(self.arg):pass#自定义方法 可以让类与类之间进行运算 对于每一种运算符 都有一种对应方法 称为运算符重载
	def function(self):pass
del instance>>>'has been deled'
instance=ClassName(arg)#类的实例化
instance.arg=value
instance.arg=ClassName2(arg)
print(ClassName.__doc__)#返回类注释
print(instance)#调用__str__()
hasattr(obj,attr)#返回对象属性存在布尔值
getattr(obj,attr,default)#返回对象属性值 如果没有属性值又没有设置默认值 会异常
setattr(obj,attr,value)#设置属性值 属性不一定要存在
delattr(obj,attr)#属性必须存在
issubclass(subclass,supclass)#如果subclass是supclass子类返回True
help(obj)#返回对象帮助信息

#collections 工具集合
import collections
#Counter 计数器
count_dict=collections.Counter(iter)#计数字典
most_list=count_dict.most_common(int)#按键出现次数从高到低排序 返回前int个键值对分别组成的元组 所构成的列表
count_list=count_dict.elements()#返回计数字典的所有键组成的列表 包含每一次
count_dict.update(dict2)#并集
count_dict.substract(dict2)#交集
count_dict.iteritems()#返回所有item
count_dict.iterkeys()#返回所有key
count_dict.itervalues()#返回所有value
for val,freq in count_dict:pass
#deque 双向队列
deque_seq=collections.deque(iter)
deque_seq.append()#队列右边添加元素
deque_seq.appendleft()#队列左边添加元素
deque_seq.clear()#清空队列中的所有元素
deque_seq.count()#返回队列中元素的个数
deque_seq.extend()#队列右边扩展 可以是列表 元组或字典 如果是字典则将字典的key加入到deque
deque_seq.extendleft()#同extend 在左边扩展
deque_seq.pop()#移除并返回队列右边的元素
deque_seq.popleft()#移除并返回队列左边的元素
deque_seq.remove(value)#移除队列第一个出现的元素
deque_seq.reverse()#队列的所有元素进行反转
deque_seq.rotate(index)#以index为中心 把所有元素索引进行旋转
#defaultdict 默认字典
default_dict=collections.defaultdict(dict)#继承普通字典 同时键值对有默认值 即使键还不存在
#OrderedDict 有序字典 继承普通字典 同时键值对有顺序
#namedtuple 可命名元组
named_tuple=collections.namedtuple(*attr)
tuple_name=named_tuple(*attr_value)#对元组命名

#openpyxl excel表格处理
import openpyxl#需要安装 pip install openpyxl
from openpyxl import Workbook
xlsx=Workbook()#创建xlsx
xlsx=openpyxl.load_workbook(filename,data_only=False)#打开xlsx文件 类的实例化 data_only=True返回单元格的计算后的值 False将返回单元格的原字符串
xlsx.sheetnames#xlsx文件所有sheet名字组成的列表 只能引用 不能命名
xlsx.save(filename.xlsx)#程序结束时使用 只支持xlsx格式
xlsx.remove(sheet_name)#删除sheet
del xlsx['sheet_name']#删除sheet
copy_sheet=xlsx.copy_worksheet(source_xlsx)#sheet副本
sheet=xlsx.create_sheet('sheet_name',index)#新建sheet 插到索引位 索引值从0开始 不填默认插到最后
sheet=xlsx.active#激活的sheet 相当于xlsx打开后默认显示的sheet
sheet=xlsx['sheet_name']#打开xlsx中对应名字的sheet
sheet.title#表格名字 只能引用 和更改 不能命名
sheet.append(list)#写入行 从左下角第一个空单元格开始向右填充值 最多嵌套两层列表 列表中的列表表示行 列表的元素表示单元格值
for row in sheet.values:#sheet所有值数据 不包含值格式
	for value in row:pass
sheet.sheet_properties.tabColor='1072BA'#改变sheet标签颜色
sheet.max_row#最大行 只能引用 不能命名
sheet.max_column#最大列 只能引用 不能命名
sheet.row_dimensions[num_index].height = 40#行高 类似列表切片 默认12.75
sheet.column_dimensions['letter_index'].width = 30#列宽 类似列表切片 默认8.43
sheet.merge_cells('A1:C3')#合并一个矩形区域中的单元格
sheet.unmerge_cells('A1:C3')#拆分一个矩形区域中的单元格 只会保留左上角值在左上角单元格
sheet.freeze_panes='coordinate'#冻结行 列 滚动sheet时 冻结的行列不会被滚动
list(sheet.rows)#所有行 每行单元格共组成一个元组 所有元组组成一个生成器
list(sheet.columns)#所有列 每列单元格共组成一个元组 所有元组组成一个生成器
a1=sheet['A1']#单元格 'ColumnRow'
a1.value#单元格值 只能引用 不能命名 如果是日期格式 会自动转为datetime.datetime()类
a1.row#单元格行值 只能引用 不能命名
a1.column#单元格列值 只能引用 不能命名
a1.coordinate#单元格坐标 只能引用 不能命名
a1=sheet.cell(row=row_num,column=col_num,value=None)#把value写入excel 用数字来代替字母形式 数字从1开始 而不是0 value填充会覆盖单元格值 不写将返回单元格原有的值
from openpyxl.utils import get_column_letter, column_index_from_string
get_column_letter(int)#返回数字对应的列的字母
column_index_from_string('letter')#返回字母对应的列的数字
a1=value#填充值 可以输入excel公式 如 "=SUM(A1, B2)"
a1=datetime.datetime.now().strftime("%Y-%m-%d")#填充当前日期
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment#单元格格式
a1.font=Font(name='等线', size=24, italic=True, underline=True,color=colors.RED, bold=True)#等线 24号 加粗 斜体 下划线 红色 默认字体11
a1.alignment=Alignment(horizontal='center', vertical='right')#水平居中 竖直居右
left, right, top, bottom = [Side(style='thin', color='000000')] * 4
a1.border = Border(left=left, right=right, top=top, bottom=bottom)#边框
cell_slice=sheet['coordinate_start':'coordinate_end']
column=sheet['column_letter']
column_slice=sheet['letter_start:letter_end']
row=sheet[row_num]
row_slice=sheet[num_start:num_end]
for row in sheet.iter_rows(min_row=row_num,min_col=col_num,max_row=row_num,max_col=col_num,values_only=False):#指定行 values_only 只有值 没有坐标
	for cell in row:pass
for col in sheet.iter_cols(min_row=row_num,min_col=col_num,max_row=row_num,max_col=col_num,values_only=False):pass#指定列 values_only 只有值 没有坐标
for row in sheet.rows:
	for cell in row:
		print(cell.value,cell.coordinate)
for column in sheet.columns:
	for cell in column:
		print(cell.value,cell.coordinate)
cols_list=list(zip(*rows_list))#矩阵置换 矩阵旋转 行转列 列转行 若某一单元格缺少数据 会被舍弃这一列/行
from openpyxl.drawing.image import Image#插入图像
sheet.add_image(Image('logo.png'), 'A1')#添加到工作表并锚定在单元格旁边

import re#正则表达式
#正则表达式常用方法
match_result=re.match(pattern,string,flags)#以pattern正则表达式从string起始位置匹配 只返回一个 无匹配结果返回None
search_result=re.search(pattern,string,flags)#以pattern正则表达式搜索string直到找到一个匹配 只返回一个 无匹配结果返回None
sub_result=re.sub(pattern,substitute,string,count,flags)#以pattern正则表达式搜索string 找到一个匹配 将匹配字符串替换为substitute 无匹配结果返回None 替换count次 默认全部替换 substitute可以是函数
split_result=re.split(pattern,string,count,flags)#以pattern正则表达式匹配string后返回列表
def multiply_2(matched):return str(int(matched.group('value')) * 2)
'A46G8HFD1134'==re.sub(r'(?P<value>\d+)', multiply_2, 'A23G4HFD567')#将字符串中的数字x2

#编译正则表达式
regex_pattern=re.compile(pattern,flags)#生成一个正则表达式
match_result=regex_pattern.match(string,start_index,end_index)#返回一个match对象
fullmatch_result=regex_pattern.fullmatch(string,start_index,end_index)#整个 string 匹配正则表达式 才返回 否则返回None
match_result.group(index)#返回匹配组 0表示所有分组组成的字符 group从1开始
match_result.start()#返回匹配结果第一个字符在string的索引
match_result.end()#返回匹配结果最后一个字符在string的索引
match_result.span()#返回(start_index,end_index)
search_result=regex_pattern.search(string,start_index,end_index)#返回一个search对象 用法同match对象
findall_result=regex_pattern.findall(string,start_index,end_index)#返回所有匹配组成的列表 匹配所有 如果没有匹配返回空列表 如果正则表达式分组 返回由分组组成的每个元组组成的列表
finditer_result=regex_pattern.finditer(string,start_index,end_index)#和findall类似 但返回迭代器
sub_result=regex_pattern.sub(substitute,string,count,flags)#将匹配字符串替换为substitute后返回string 无匹配结果返回原string 替换count次 默认全部替换 substitute可以是函数
subn_result=regex_pattern.subn(substitute,string,count,flags)#和 sub类似 但返回元组 (string,count)
escape_result=regex_pattern.escape(pattern)#转义pattern中具有正则表达式特殊含义的字符
'|'.join(map(re.escape, sorted(['+', '-', '*', '/', '**'], reverse=True)))==r'/|\-|\+|\*\*|\*'

#flags控制匹配方式
re.I#re.IGNORECASE 忽略字母大小写
re.A#re.ASCII 只匹配ASCII字符
re.DEBUG#显示编译时的debug信息
re.L#re.LOCALE 基本没用 只能对byte样式有效 由当前语言区域决定 \w, \W, \b, \B 和大小写敏感匹配
re.M#re.MULTILINE 使^除了匹配整个字符串的起始位置，还匹配换行符\n后面的位置 $除了匹配整个字符串的结束位置，还匹配换行符\n前面的位置
re.S#re.DOTALL 使 . 能够匹配所有字符
re.X#re.VERBOSE 忽略表达式中的空白 和 #注释 让长正则表达式易于理解
verbose_regex = re.compile(
r'''( (\d{3}|\(\d{3}\))?	# area code
(\s|-|\.)?					# separator
\d{3}						# first 3 digits
(\s|-|\.)						# separator
\d{4}						# last 4 digits
)''',re.X)
re.findall(r'^This.*?line.$', 'This is the first line.\nThis is the second line.\nThis is the third line.', flags=re.M+re.S)==['This is the first line.', 'This is the second line.', 'This is the third line.']#匹配多行文本的每一行
flags=re.M | re.S#另一种flags表示
re.findall(r'(?ms)^This.*?line.$')#另一种flags表示 这种表达方式不需要写flags 但必须在表达式开头用(?)声明 可选参数有(?aiLmsux)

#(?)参数
(str)#分组 索引从1开始 0表示所有组组成的字符串
(?aiLmsx)#字母分别代表 re.A I L M S X 必须在表达式开头使用
(?:str)#在使用group之类的函数时忽略该组
(?P<group_name>str)#为该组命名 使用groupdict()时键为group_name 值为匹配的组
(?P=group_name)#引用(?P<group_name>str)
(?#str)#str会被忽略 当作注释
(?=str)#只有匹配到str的时候 才会匹配(?=str)前面的表达式
(?!str)#只有没有匹配到str的时候 才会匹配(?!str)前面的表达式
(?<=str)#在字符串非开始位置匹配以str开头的字符串 只能是明确的长度 可以理解成 ^str 在字符串非开始位置匹配符合字符
(?<!str)#在字符串非开始位置匹配不以str开头的字符串 只能是明确的长度 可以理解成 ^str 在字符串非开始位置匹配不符合字符
(?(group_index/group_name)yes-pattern|no-pattern)#如果给定的 group_index 或 group_name 存在，将会尝试匹配 yes-pattern ，否则就尝试匹配 no-pattern，no-pattern 可选，也可以被忽略。比如，
r'(<)?(\w+@\w+(?:\.\w+)+)(?(1)>|$)'#是一个email样式匹配，将匹配 '<user@host.com>' 或 'user@host.com' ，但不会匹配 '<user@host.com' ，也不会匹配 'user@host.com>'
match_result.group()#返回匹配的整个结果
match_result.group(int)#返回匹配结果的第int组 从1开始
match_result.groups()#返回匹配的所有组
match_result.groupdict()#返回匹配的所有组名和组组成的字典 键为组名 值为匹配的组

#匹配模式
r'pattern'#r表示忽略转义字符 写正则表达式加上就行
贪婪模式#有多种匹配字符串符合正则表达式 匹配字符串最长的 默认模式
非贪婪模式#有多种匹配字符串符合正则表达式 匹配字符串最短的 用?在正则式结尾修饰 如 .*?
'^str'#只匹配以str开头的字符串 caret ^必须在$之前 caret before dollar ==abcd
'str$'#只匹配以str结尾的字符串 dollar ^必须在$之前 caret before dollar ==abcd
.#匹配任意一个除了\n的字符 当re.M标记被指定时 则可以匹配包括换行符的任意字符
[str]#匹配在[]的一个字符 如[amk] 匹配 'a' 'm'或'k'
[^str]#不在[]中的字符：[^abc] 匹配除了a,b,c之外的字符
(str)*#匹配0个到n个str 贪婪模式
(str)*?#匹配0个到n个str 非贪婪模式
(str)+#匹配1个到n个str 贪婪模式
(str)+?#匹配1个到n个str 非贪婪模式
(str)?#匹配时str可有可无 贪婪模式
(str)??#匹配时str可有可无 非贪婪模式 即匹配结果为空 ""
(str){ x}#只匹配重复x次的str 例如  (o){2} 不能匹配 "Bob" 中的 "o" 但是能匹配 "food" 中的两个 o
(str){ x,}#只匹配重复x次到n次的str 例如  o{2,} 不能匹配"Bob"中的"o" 但能匹配 "foooood"中的所有 o        "o{1,}" 等价于 "o+"            "o{0,}" 则等价于 "o*"
(str){ ,y}#只匹配重复0次到y次的str 例如  o{2,} 不能匹配"Bob"中的"o" 但能匹配 "foooood"中的所有 o        "o{1,}" 等价于 "o+"            "o{0,}" 则等价于 "o*"
(str){ x, y}#只匹配重复x到y次的str 贪婪模式
(str){ x, y}?#只匹配重复x到y次的str 非贪婪模式 或 str可有可无
a|b#匹配a或b

r'\w' #匹配一个字母数字及下划线 [A-Za-z0-9_] word：3rd-Person
r'\W'#匹配一个非字母数字及下划线 '[^A-Za-z0-9_]'
r'\s'#匹配一个任意空白字符 等价于  [ \f\n\r\t\v]
r'\S'#匹配一个任意非空字符 [^ \f\n\r\t\v]
r'\d'#匹配一个任意数字 等价于 [0-9].
r'\D'#匹配一个任意非数字 [^0-9]
r'\A'#类似^ 但不受re.M控制
r'\Z'#类似$ 但不受re.M控制
r'\z'#匹配一个字符串结束
r'\G'#匹配一个最后匹配完成的位置
'\b'#匹配一个单词边界 如 r'\bfoo\b' 匹配 'foo', 'foo.', '(foo)', 'bar foo baz' 但不匹配 'foobar' 或者 'foo3'
r'\B'#匹配非单词边界 如 r'py\B' 匹配 'python', 'py3', 'py2', 但不匹配 'py', 'py.', 或者 'py!'. \B 是 \b 的取非
'\n'#匹配一个换行符
'\t'#匹配一个制表符
'\1'-'\99'#第1-99个分组的内容 表达式自身也可以引用
re.compile(r'Agent (\w)\w*').sub(r'\1****', 'Agent Alice told Agent Carol that Agent Eve knew Agent Bob was a spy.')=='A**** told C**** that E**** knew B**** was a spy.'

[0-9]#匹配任何数字 类似于 [0123456789]
[a-z]	#匹配任何小写字母
[A-Z]#匹配任何大写字母
[a-zA-Z0-9]#匹配任何字母及数字
[^aeiou]#除了aeiou字母以外的所有字符
[^0-9]#匹配除了数字外的字符

r'\d(.*) are (.*?) .*'
r 表示忽略转义字符
(.*) 第一个匹配分组，.* 代表匹配除之外的所有长度的字符
(.*?) 第二个匹配分组，.*? 后面多个问号，代表非贪婪模式，只匹配符合条件的最少字符
re.compile(r'<.*?>').search('<To serve man> for dinner.>').group()=='<To serve man>'
re.compile(r'<.*>').search('<To serve man> for dinner.>').group()=='<To serve man> for dinner.>'

#常用正则表达式 https://www.cnblogs.com/magicking/p/8986869.html

pyInstaller #python打包 .exe程序 在别的电脑有几率exe出错 原因是别的电脑缺少相关运行库 过几年等pyInstaller自己更新打包算法
#安装打包模块之前 先升级pip
#GUI开发选用tkinter会减少很多麻烦事

简单程序打包
.py文件储存目录 打开cmd输入以下任意代码
#用于简单测试程序 单个.exe文件 保留打包过程产生的日志和缓存
pyinstaller -F main_script.py
#-F 只生成一个exe文件 如果有多个.py文件 或自带资源 不能带该参数 文件复杂不建议使用该参数
#最终程序 目录形式 不含终端 含程序图标删除打包过程产生的日志和缓存
pyinstaller -D -w -i icon.ico --clean main_script.py
pyinstaller -F -w -i icon.ico --clean main_script.py
#-D 打包成一个目录的形式 为打包的默认形式 可以不加
#-w 执行.exe文件时隐藏控制台 建议只有在已经设置了GUI时才带这个参数 只对windows有效
#-c 执行.exe文件时显示控制台 默认会显示 可以不加
#-i 替换程序图标
#icon.ico 程序同级目录里的程序图标
#--clean 在打包前清除pyinstaller缓存
#--debug=all 显示所有程序诊断信息
#main_script.py 将要打包的程序 如果有多个py文件 main_script是程序的主程序
#打包完成后在py目录中 spec和build是打包过程产生的数据和日志文件 pycache是缓存文件 dist是程序文件
#打包结束除了dist文件都可以删除 dist是.exe所在目录 如果程序使用了自带资源 要把自带资源放在.exe同级目录 也就是dist内 dist可以重命名
例如使用了icon.ico 需要放在exe同级目录

复杂程序打包
pip install pipenv#cmd输入 安装pip虚拟环境 直接打包 程序会很大 在虚拟环境进行打包 如果对文件大小不在意 可以直接跳过虚拟环境设置部分

选择一个目录作为虚拟环境目录 在该目录下 cmd 创建虚拟环境
pipenv install --python 3.9.5
#选择的目录是 Python_Virtual_Env python版本要符合使用的版本
#pipenv安装packages会生成一个Pipfile.lock的文件 这个文件主要是关于所下载的packages的一些信息
打开虚拟环境中的Pipfile
修改url = "https://pypi.tuna.tsinghua.edu.cn/simple/"#包下载地址替换为清华源
pipenv shell#cmd 激活虚拟环境 每次安装包都要先输入
exit#cmd 退出虚拟环境
#若package安装失败 将Pipfile.lock删除 然后在控制台中通过pipenv lock来生成新的Pipfile.lock
在虚拟环境下安装需要打包的程序的依赖的库#模块之间用空格分开
pip install pywin32 pyinstaller
在虚拟环境下进行打包

在程序同级目录 打开cmd输入
pyi-makespec -w main_script.py
会生成一个对于 main_script.py的 main_script.spec文件#main_script.py 是将要打包的主程序文件
#spec文件中主要包含 Analysis PYZ EXE COLLECT
#Analysis以py文件为输入 它会分析py文件的依赖模块 并生成相应的信息
#PYZ是一个.pyz的压缩包 包含程序运行需要的所有依赖
#EXE根据上面两项生成
#COLLECT生成其他部分的输出文件夹 COLLECT也可以没有
#官方spec说明 https://pyinstaller.readthedocs.io/en/v4.5.1/spec-files.html#using-spec-files
打开 .spec文件 依照下列代码进行内容修改 注释可以不用删
# -*- mode: python ; coding: utf-8 -*-
#指定utf-8编码

#出错 提示RecursionError: maximum recursion depth exceeded 用下面两行代码 一般可以不管 5000可以替换成其他数值
#import sys
#sys.setrecursionlimit(5000)

block_cipher = None

a = Analysis(['D:\\script_folder1\\script_folder2\\main_script.py',
	'D:\\script_folder1\\script_folder2\\aux_script1.py',
	'D:\\script_folder1\\script_folder2\\aux_script1.py'],
#修改为需要打包的.py文件 完整路径 程序所有相关的.py文件都要写入 注意必须是\\
	pathex=['D:\\script_folder1\\script_folder2',
	'D:\\python39\\Lib\\site-packages\\used_package_folder1',
	'D:\\python39\\Lib\\site-packages\\used_package_folder2'],
#修改为.py文件import的模块所在文件夹 包括python模块 三方模块 自己写的模块 site-packages文件夹里的才是三方模块 注意必须是\\ 完整路径 如果程序运行出错 提示缺少什么就补相应的文件夹
	binaries=[('C:\\Windows\\SysWOW64\\api-ms-win-core-path-l1-1-0.dll','.')],#所有二进制资源储存的文件夹或文件 如dll 动态库 一般不用管 为空列表 列表元素格式和datas一样
	#打包后的exe在没有Python环境上运行可能提示缺少api-ms-win-core-path-l1-1-0.dll
	#参考 http://www.xitongzhijia.net/soft/201409.html 下载该dll
	#将该dll复制一份到Python根目录
	#将该dll复制一份到Windows\SysWOW64
	#将该dll复制一份到Windows\System32\downlevel
	datas=[('D:\\script_folder1\\script_folder2\\images\\role_images','images\\role_images'),
	('D:\\script_folder1\\script_folder2\\sound','sound'),
	('D:\\script_folder1\\script_folder2\\*.txt','.'),
	('D:\\script_folder1\\script_folder2\\icon.ico','.')],#所有资源储存的文件夹或文件 如程序内用到的exe图标 声音 图片 数据文件等 元组格式为(资源文件储存路径,打包后相对exe的路径) . 表示exe所在目录 *表示匹配任意多个字符
	hiddenimports=['module_name1','module_name2'],#打包后执行程序时出现类似No Module named xxx 在此添加程序中引入的模块名称 一般设置为空列表 []
	hookspath=[],
	hooksconfig={},
	runtime_hooks=[],
	excludes=['module_name1','module_name2'],#打包时不打包的模块 一般设置为空列表 []
	win_no_prefer_redirects=False,
	win_private_assemblies=False,
	cipher=block_cipher,
	noarchive=False)
pyz = PYZ(a.pure, a.zipped_data,
	cipher=block_cipher)
exe = EXE(pyz,
	a.scripts,
	a.binaries,
	a.zipfiles,
	a.datas,
	[],
	exclude_binaries=True,
	name='exe_name',
	#name为.exe名
	debug=False,#调试测bug的时候可以改为True
	bootloader_ignore_signals=False,
	strip=False,
	upx=True,
	upx_exclude=[],
	runtime_tmpdir=None,
	console=False,#True显示控制台 False不显示
	disable_windowed_traceback=False,
	target_arch=None,
	codesign_identity=None,
	entitlements_file=None,
	icon='D:\\script_folder1\\script_folder2\\main_script.ico'#exe程序图标 必须是ico格式 完整路径 注意是\\
	)
coll = COLLECT(exe,
	a.binaries,
	a.zipfiles,
	a.datas,
	strip=False,
	upx=True,
	upx_exclude=[],
	name='exe_name')#name为.exe名

修改完成后 cmd输入
pyinstaller -D --clean main_script.spec
#生成的build文件可以删除 __pycache__ .spec最好不删 exe在dist内 若有自定义资源 或三方库 要复制到和exe同级目录

#打包后的程序还是很大 用upx对exe文件进行压缩 https://upx.github.io upx
在程序目录 cmd输入#upx_file是upx储存目录
pyinstaller --upx-dir upx_file --clean main.py

if __name__ == '__main__':
    multiprocessing.freeze_support()#如果调用多线程 必须这么写 否则运行exe会显示 Failed to execute script pyi_rth_multiprocessing

#以下备注不确定是否有效
#
if __name__ == '__main__':#需要打包的程序不要使用该代码 否则可能出错
#需要打包的程序 使用from..import... 减少程序大小
#程序中使用的三方库要复制到程序同目录 否则可能出错
#

#Python>=3.5开发的程序 在非Windows10使用 程序可能会出现 Visual C++ run-time .dlls的错误 若遇到此问题 解决方法之一是在win7上开发程序 其他方法略
#更多见 https://blog.csdn.net/weixin_42052836/article/details/82315118?utm_medium=distribute.pc_relevant.none-task-blog-2%7Edefault%7EBlogCommendFromMachineLearnPai2%7Edefault-1.essearch_pc_relevant&depth_1-utm_source=distribute.pc_relevant.none-task-blog-2%7Edefault%7EBlogCommendFromMachineLearnPai2%7Edefault-1.essearch_pc_relevant

cx_freeze#另一种打包模块
#https://www.jianshu.com/p/c029574187d2
